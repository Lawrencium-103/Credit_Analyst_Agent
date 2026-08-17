"""Generic xlsx parser (non-Standard-Chartered statements).

Best-effort: scans every sheet for known labels and reads ALL numeric data
columns. Handles the messy Excel files that junior credit analysts actually
receive:

  - Merged cells (auto-unmerged during read)
  - Hidden rows/columns (skipped with flag)
  - Vertical layouts (years as rows, labels as columns)
  - Multi-sheet workbooks (auto-detect IS/BS/CF sheets)
  - Different number formats (parentheses, currency, percentages)
  - Different row label variations (60+ synonyms)

Output is flagged low-confidence for human review.
"""

from __future__ import annotations

import re

import openpyxl
from openpyxl.cell.cell import MergedCell

from ..schema.financials import BalanceSheet, CashFlow, IncomeStatement, PeriodFinancials
from .number_parser import parse_number

# ── Extended keyword list (field → ordered list of substring matches) ────────
_KEYWORDS: list[tuple[str, list[str]]] = [
    # Income statement
    ("revenue", [
        "total revenue", "net revenue", "total net revenue",
        "net sales", "revenue from operations", "turnover",
        "revenue", "sales",
    ]),
    ("cogs", [
        "cost of goods sold", "cost of products sold",
        "cost of sales", "cost of revenue", "cogs",
    ]),
    ("gross_profit", ["gross profit", "gross income"]),
    ("operating_expenses", [
        "total operating expenses", "selling general and administrative",
        "selling, general and administrative", "selling & administrative",
        "operating expenses", "total opex", "opex",
        "administrative expenses", "selling expenses",
    ]),
    ("ebitda", ["ebitda"]),
    ("depreciation_amortization", [
        "depreciation and amortisation", "depreciation and amortization",
        "depreciation & amortisation", "depreciation & amortization",
        "depreciation, amortisation and impairment",
        "depreciation", "amortisation", "amortization",
    ]),
    ("ebit", [
        "operating profit", "operating income",
        "income from operations", "profit from operations",
        "earnings before interest and tax", "ebit",
    ]),
    ("interest_expense", [
        "finance costs", "finance expense", "interest expense",
        "interest and finance costs", "net finance costs",
    ]),
    ("interest_income", [
        "interest income", "interest earned", "finance income",
    ]),
    ("pretax_income", [
        "profit before taxation", "profit before tax",
        "income before income taxes", "income before tax",
        "earnings before tax", "pretax income", "pbt",
    ]),
    ("tax_expense", [
        "income tax expense", "tax expense",
        "taxation", "income taxes",
    ]),
    ("net_income", [
        "net profit", "net income", "profit for the period",
        "profit attributable to", "profit after tax",
        "earnings per share", "pat", "net earnings",
    ]),
    # Balance sheet
    ("cash_and_equivalents", [
        "cash and cash equivalents", "cash at bank",
        "cash and short-term deposits",
    ]),
    ("marketable_securities", [
        "short-term investments", "marketable securities",
        "trading securities", "financial assets at fair value",
    ]),
    ("accounts_receivable", [
        "trade receivables", "accounts receivable",
        "receivables, net", "trade and other receivables",
    ]),
    ("inventory", ["inventories", "inventory"]),
    ("current_assets", ["total current assets", "current assets"]),
    ("total_assets", ["total assets"]),
    ("accounts_payable", [
        "trade payables", "accounts payable",
        "trade and other payables", "payables",
    ]),
    ("current_liabilities", [
        "total current liabilities", "current liabilities",
    ]),
    ("total_liabilities", ["total liabilities"]),
    ("short_term_debt", [
        "short-term borrowings", "current portion of long-term debt",
        "short-term debt", "current maturities of debt",
        "short-term borrowings and current",
    ]),
    ("long_term_debt", [
        "long-term borrowings", "long-term debt",
        "non-current liabilities", "long-term borrowings and lease",
    ]),
    ("total_debt", [
        "total borrowings", "total debt",
        "total financial liabilities",
    ]),
    ("total_equity", [
        "total shareholders equity", "shareholders equity",
        "stockholders equity", "total equity",
        "equity attributable to", "total stockholders",
    ]),
    ("retained_earnings", [
        "retained earnings", "accumulated deficit",
        "retained earnings (accumulated deficit)",
    ]),
    # Cash flow
    ("operating_cash_flow", [
        "net cash from operating activities",
        "cash generated from operations",
        "cash flow from operating activities",
        "net cash provided by operating activities",
        "operating cash flow",
    ]),
    ("capital_expenditures", [
        "purchase of property and equipment",
        "purchases of property and equipment",
        "acquisition of property, plant",
        "additions to property, plant",
        "capital expenditure", "capex",
    ]),
    ("free_cash_flow", ["free cash flow", "free cashflow"]),
    ("investing_cash_flow", [
        "net cash used in investing activities",
        "cash flow from investing activities",
        "investing cash flow",
    ]),
    ("financing_cash_flow", [
        "net cash used in financing activities",
        "cash flow from financing activities",
        "financing cash flow",
    ]),
    ("dividends_paid", [
        "dividends paid", "dividends paid to shareholders",
        "dividends declared",
    ]),
]

# Keywords for detecting which sheet contains which statement
_IS_KEYWORDS = ["income", "profit", "loss", "p&l", "pnl", "statement of earnings",
                "statement of operations", "operating"]
_BS_KEYWORDS = ["balance", "position", "statement of financial"]
_CF_KEYWORDS = ["cash flow", "cashflow", "cashflows", "funds flow"]

# Year-like patterns for header detection
_YEAR_RE = re.compile(r"(?:FY\s*)?(?:20)?(\d{2})(?:\s*[/\-\u2013]\s*(?:FY\s*)?(?:20)?(\d{2}))?")
_DATE_RE = re.compile(r"\d{4}[-/]\d{1,2}(?:[-/]\d{1,2})?")
_FULL_YEAR_RE = re.compile(r"(?:19|20)\d{2}")


# ── Merged cell handling ─────────────────────────────────────────────────────

def _unmerge_and_fill(ws) -> None:
    """Unmerge all merged ranges and fill every cell with the top-left value.

    This allows the parser to read data from cells that are visually merged
    but technically empty in openpyxl.
    """
    merges = list(ws.merged_cells.ranges)
    for rng in merges:
        top_left_value = ws.cell(row=rng.min_row, column=rng.min_col).value
        ws.unmerge_cells(str(rng))
        for row in range(rng.min_row, rng.max_row + 1):
            for col in range(rng.min_col, rng.max_col + 1):
                ws.cell(row=row, column=col).value = top_left_value


# ── Hidden row/column detection ──────────────────────────────────────────────

def _count_hidden(ws) -> tuple[int, int]:
    """Return (hidden_rows, hidden_columns) counts."""
    hidden_rows = sum(
        1 for r in range(1, (ws.max_row or 0) + 1)
        if ws.row_dimensions[r].hidden
    )
    hidden_cols = sum(
        1 for c_letter in _col_letters(ws)
        if ws.column_dimensions[c_letter].hidden
    )
    return hidden_rows, hidden_cols


def _col_letters(ws) -> list[str]:
    """Get all column letters that have dimensions set."""
    return [k for k in ws.column_dimensions if isinstance(k, str) and len(k) <= 3]


def _is_row_hidden(ws, row: int) -> bool:
    return ws.row_dimensions[row].hidden


def _is_col_hidden(ws, col: int) -> bool:
    from openpyxl.utils import get_column_letter
    letter = get_column_letter(col)
    return ws.column_dimensions[letter].hidden


# ── Year/period header detection ─────────────────────────────────────────────

def _is_period_header(val) -> bool:
    """Return True if a cell value looks like a period header (year/date).

    The ENTIRE value must match — substrings of larger numbers are rejected.
    """
    if val is None:
        return False
    s = str(val).strip()
    if not s:
        return False
    if _FULL_YEAR_RE.fullmatch(s):
        return True
    if _DATE_RE.fullmatch(s):
        return True
    if re.fullmatch(r"(?:FY\s*)?(?:20)?\d{2}", s, re.IGNORECASE):
        return True
    if re.fullmatch(r"(?:20)?\d{2}\s*[/\-\u2013]\s*(?:20)?\d{2}", s):
        return True
    return False


def _extract_year_label(val) -> str | None:
    """Extract a clean year label from a period header cell."""
    if val is None:
        return None
    s = str(val).strip()

    if _FULL_YEAR_RE.fullmatch(s):
        return s[:4]

    m = _DATE_RE.fullmatch(s)
    if m:
        ym = _FULL_YEAR_RE.search(s)
        if ym:
            return ym.group(0)[:4]

    m = re.fullmatch(r"(?:FY\s*)?(?:20)?(\d{2})", s, re.IGNORECASE)
    if m:
        y1 = int(m.group(1))
        yr = 2000 + y1 if y1 < 100 else y1
        return str(yr)

    m = re.fullmatch(r"(?:20)?(\d{2})\s*[/\-\u2013]\s*(?:20)?(\d{2})", s)
    if m:
        y2 = int(m.group(2))
        yr2 = 2000 + y2 if y2 < 100 else y2
        return str(yr2)

    return None


# ── Layout detection ─────────────────────────────────────────────────────────

def _detect_layout(ws) -> str:
    """Detect whether the data is laid out horizontally or vertically.

    Horizontal: labels in column A, data in columns B, C, D...
    Vertical:   labels in row 1, data in rows 2, 3, 4...  (years as rows)

    Returns "horizontal" or "vertical".
    """
    if ws.max_row is None or ws.max_column is None:
        return "horizontal"

    # Signal 1: Column A has multiple period-like values (years/dates) in rows 2+
    period_in_col_a = 0
    for row in range(2, min((ws.max_row or 1) + 1, 50)):
        val = ws.cell(row=row, column=1).value
        if val is not None and _is_period_header(val):
            period_in_col_a += 1

    # Signal 2: Row 1 has text labels in columns B+ (not numeric, not empty)
    text_in_row_1 = 0
    for col in range(2, min((ws.max_column or 1) + 1, 20)):
        val = ws.cell(row=1, column=col).value
        if val is not None and not isinstance(val, (int, float)):
            text_in_row_1 += 1

    # Vertical if: col A has multiple year labels AND row 1 has text column headers
    if period_in_col_a >= 2 and text_in_row_1 >= 2:
        return "vertical"

    return "horizontal"


# ── Multi-sheet IS/BS/CF detection ───────────────────────────────────────────

def _classify_sheet(sheet_name: str) -> str | None:
    """Classify a sheet name as income_statement, balance_sheet, cash_flow, or None."""
    name_l = sheet_name.lower()
    for kw in _CF_KEYWORDS:
        if kw in name_l:
            return "cash_flow"
    for kw in _BS_KEYWORDS:
        if kw in name_l:
            return "balance_sheet"
    for kw in _IS_KEYWORDS:
        if kw in name_l:
            return "income_statement"
    return None


def _has_is_labels(ws) -> float:
    """Check if a worksheet contains income statement labels. Returns ratio."""
    count = 0
    checked = 0
    for row in range(1, min(ws.max_row or 1, 50) + 1):
        label = ws.cell(row=row, column=1).value
        if not label:
            continue
        checked += 1
        label_l = str(label).lower()
        for field, kws in _KEYWORDS:
            if field not in IncomeStatement.model_fields:
                continue
            for kw in kws:
                if kw in label_l:
                    count += 1
                    break
    return count / max(checked, 1)


def _has_bs_labels(ws) -> float:
    """Check if a worksheet contains balance sheet labels. Returns ratio."""
    count = 0
    checked = 0
    for row in range(1, min(ws.max_row or 1, 50) + 1):
        label = ws.cell(row=row, column=1).value
        if not label:
            continue
        checked += 1
        label_l = str(label).lower()
        for field, kws in _KEYWORDS:
            if field not in BalanceSheet.model_fields:
                continue
            for kw in kws:
                if kw in label_l:
                    count += 1
                    break
    return count / max(checked, 1)


def _has_cf_labels(ws) -> float:
    """Check if a worksheet contains cash flow labels. Returns ratio."""
    count = 0
    checked = 0
    for row in range(1, min(ws.max_row or 1, 50) + 1):
        label = ws.cell(row=row, column=1).value
        if not label:
            continue
        checked += 1
        label_l = str(label).lower()
        for field, kws in _KEYWORDS:
            if field not in CashFlow.model_fields:
                continue
            for kw in kws:
                if kw in label_l:
                    count += 1
                    break
    return count / max(checked, 1)


# ── Column detection ─────────────────────────────────────────────────────────

def _detect_data_columns(ws, skip_hidden: bool = True) -> list[int]:
    """Detect which columns contain numeric data."""
    numeric_cols: dict[int, int] = {}
    sample_rows = min(ws.max_row or 1, 30)

    for row in range(1, sample_rows + 1):
        if skip_hidden and _is_row_hidden(ws, row):
            continue
        for col in range(2, (ws.max_column or 2) + 1):
            if skip_hidden and _is_col_hidden(ws, col):
                continue
            val = ws.cell(row=row, column=col).value
            if val is None:
                continue
            parsed = parse_number(val)
            if parsed is not None:
                numeric_cols[col] = numeric_cols.get(col, 0) + 1

    return [col for col, count in sorted(numeric_cols.items()) if count >= 1]


def _detect_data_rows(ws, skip_hidden: bool = True) -> list[int]:
    """For vertical layouts: detect which rows contain numeric data (in column 1+)."""
    numeric_rows: dict[int, int] = {}
    sample_cols = min(ws.max_column or 1, 15)

    for col in range(2, sample_cols + 1):
        for row in range(1, (ws.max_row or 1) + 1):
            if skip_hidden and _is_row_hidden(ws, row):
                continue
            val = ws.cell(row=row, column=col).value
            if val is None:
                continue
            parsed = parse_number(val)
            if parsed is not None:
                numeric_rows[row] = numeric_rows.get(row, 0) + 1

    return [r for r, count in sorted(numeric_rows.items()) if count >= 1]


def _detect_period_headers(ws, data_cols: list[int]) -> dict[int, str]:
    """Scan the first few rows to find period labels for each data column."""
    for row in range(1, min(6, (ws.max_row or 1) + 1)):
        candidates: dict[int, str] = {}
        for col in data_cols:
            val = ws.cell(row=row, column=col).value
            year = _extract_year_label(val)
            if year:
                candidates[col] = year
        if len(candidates) >= max(1, len(data_cols) // 2):
            return candidates
    return {}


def _detect_period_headers_vertical(ws, data_rows: list[int]) -> dict[int, str]:
    """For vertical layouts: scan the first few columns to find period labels."""
    for col in range(1, min(6, (ws.max_column or 1) + 1)):
        candidates: dict[int, str] = {}
        for row in data_rows:
            val = ws.cell(row=row, column=col).value
            year = _extract_year_label(val)
            if year:
                candidates[row] = year
        if len(candidates) >= max(1, len(data_rows) // 2):
            return candidates
    return {}


# ── Label matching ───────────────────────────────────────────────────────────

def _match_field(label: str) -> str | None:
    """Match a row label to a canonical field name using longest-match-first."""
    label_l = label.lower()
    best_field = None
    best_len = 0
    for field, keywords in _KEYWORDS:
        for kw in keywords:
            if kw in label_l and len(kw) > best_len:
                best_field = field
                best_len = len(kw)
    return best_field


# ── Period construction ──────────────────────────────────────────────────────

def _to_period_dict(
    found: dict[str, float | None],
) -> tuple[dict[str, float | None], dict[str, float | None], dict[str, float | None]]:
    """Split a flat field dict into IS / BS / CF dicts."""
    is_fields = {k: v for k, v in found.items() if k in IncomeStatement.model_fields}
    bs_fields = {k: v for k, v in found.items() if k in BalanceSheet.model_fields}
    cf_fields = {k: v for k, v in found.items() if k in CashFlow.model_fields}
    return is_fields, bs_fields, cf_fields


def _build_periods(
    all_period_data: dict[str, dict[str, float | None]],
    year: str,
) -> list[PeriodFinancials]:
    """Build PeriodFinancials objects from accumulated period data."""
    periods: list[PeriodFinancials] = []
    for yr in sorted(all_period_data.keys()):
        data = all_period_data[yr]
        is_f, bs_f, cf_f = _to_period_dict(data)
        periods.append(PeriodFinancials(
            period=str(yr),
            income_statement=IncomeStatement(**is_f),
            balance_sheet=BalanceSheet(**bs_f),
            cash_flow=CashFlow(**cf_f),
        ))
    if not periods:
        periods = [PeriodFinancials(
            period=str(year),
            income_statement=IncomeStatement(),
            balance_sheet=BalanceSheet(),
            cash_flow=CashFlow(),
        )]
    return periods


# ── Horizontal layout parser ─────────────────────────────────────────────────

def _parse_horizontal(
    ws,
    base_year: str,
    all_period_data: dict[str, dict[str, float | None]],
    skip_hidden: bool = True,
) -> None:
    """Parse a horizontal layout worksheet (labels in col A, data in cols B+)."""
    data_cols = _detect_data_columns(ws, skip_hidden=skip_hidden)
    if not data_cols:
        return

    headers = _detect_period_headers(ws, data_cols)

    col_year: dict[int, str] = {}
    for col in data_cols:
        if col in headers:
            col_year[col] = headers[col]
        else:
            idx = data_cols.index(col)
            try:
                base = int(base_year)
                col_year[col] = str(base - (len(data_cols) - 1) + idx)
            except (ValueError, TypeError):
                col_year[col] = base_year

    for yr in col_year.values():
        if yr not in all_period_data:
            all_period_data[yr] = {}

    for row in range(1, (ws.max_row or 0) + 1):
        if skip_hidden and _is_row_hidden(ws, row):
            continue

        label_cell = ws.cell(row=row, column=1).value
        if not label_cell:
            continue

        field = _match_field(str(label_cell))
        if not field:
            continue

        for col in data_cols:
            yr = col_year[col]
            if field in all_period_data[yr] and all_period_data[yr][field] is not None:
                continue
            val = parse_number(ws.cell(row=row, column=col).value)
            if val is not None:
                all_period_data[yr][field] = val


# ── Vertical layout parser ───────────────────────────────────────────────────

def _parse_vertical(
    ws,
    base_year: str,
    all_period_data: dict[str, dict[str, float | None]],
    skip_hidden: bool = True,
) -> None:
    """Parse a vertical layout worksheet (labels in row 1, years as rows)."""
    data_rows = _detect_data_rows(ws, skip_hidden=skip_hidden)
    if not data_rows:
        return

    headers = _detect_period_headers_vertical(ws, data_rows)

    row_year: dict[int, str] = {}
    for r in data_rows:
        if r in headers:
            row_year[r] = headers[r]
        else:
            idx = data_rows.index(r)
            try:
                base = int(base_year)
                row_year[r] = str(base - (len(data_rows) - 1) + idx)
            except (ValueError, TypeError):
                row_year[r] = base_year

    for yr in row_year.values():
        if yr not in all_period_data:
            all_period_data[yr] = {}

    # Labels are in row 1 (columns 2+)
    for col in range(2, (ws.max_column or 1) + 1):
        if skip_hidden and _is_col_hidden(ws, col):
            continue

        label_cell = ws.cell(row=1, column=col).value
        if not label_cell:
            continue

        field = _match_field(str(label_cell))
        if not field:
            continue

        for r in data_rows:
            yr = row_year[r]
            if field in all_period_data[yr] and all_period_data[yr][field] is not None:
                continue
            val = parse_number(ws.cell(row=r, column=col).value)
            if val is not None:
                all_period_data[yr][field] = val


# ── Public API ───────────────────────────────────────────────────────────────

def parse_generic_xlsx(path: str, year: str, entity: str | None = None) -> PeriodFinancials:
    """Parse a generic xlsx into a single PeriodFinancials (backward compat)."""
    periods = parse_generic_xlsx_multi(path, year, entity)
    return periods[0] if periods else PeriodFinancials(
        period=str(year),
        income_statement=IncomeStatement(),
        balance_sheet=BalanceSheet(),
        cash_flow=CashFlow(),
    )


def parse_generic_xlsx_multi(
    path: str, year: str, entity: str | None = None
) -> list[PeriodFinancials]:
    """Parse a generic xlsx into one or more PeriodFinancials objects.

    Handles:
    - Horizontal layouts (labels in col A, years in cols B+)
    - Vertical layouts (labels in row 1, years as rows)
    - Merged cells (auto-unmerged)
    - Hidden rows/columns (skipped)
    - Multi-sheet workbooks (scans all sheets)
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    all_period_data: dict[str, dict[str, float | None]] = {}

    for ws in wb.worksheets:
        if ws.max_row is None or ws.max_row < 1:
            continue

        # Unmerge cells so merged labels/values are accessible
        _unmerge_and_fill(ws)

        # Detect layout orientation
        layout = _detect_layout(ws)

        if layout == "vertical":
            _parse_vertical(ws, year, all_period_data)
        else:
            _parse_horizontal(ws, year, all_period_data)

    wb.close()
    return _build_periods(all_period_data, year)


def describe_workbook(path: str) -> dict:
    """Return metadata about an xlsx file for the SPA to display.

    Returns dict with:
      - sheet_names: list of sheet names
      - sheet_types: {name: "income_statement"|"balance_sheet"|"cash_flow"|None}
      - hidden_rows: {sheet_name: count}
      - hidden_cols: {sheet_name: count}
      - merged_cells: {sheet_name: count}
      - detected_layout: {sheet_name: "horizontal"|"vertical"}
    """
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    info: dict = {
        "sheet_names": wb.sheetnames,
        "sheet_types": {},
        "hidden_rows": {},
        "hidden_cols": {},
        "merged_cells": {},
        "detected_layout": {},
    }
    wb.close()

    # Re-open without read_only to access dimensions
    wb2 = openpyxl.load_workbook(path, data_only=True)
    for name in wb2.sheetnames:
        ws = wb2[name]
        info["sheet_types"][name] = _classify_sheet(name)
        hr, hc = _count_hidden(ws)
        info["hidden_rows"][name] = hr
        info["hidden_cols"][name] = hc
        info["merged_cells"][name] = len(ws.merged_cells.ranges)
        info["detected_layout"][name] = _detect_layout(ws)
    wb2.close()

    return info
