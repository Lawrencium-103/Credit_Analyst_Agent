"""Spreading layer: parse the Standard Chartered example-answer workbook.

This loader maps the fixed-layout SC financial reporting workbook into the
canonical schema. It is the reference implementation of the spreading step: raw
disclosed line items are normalised (signs, magnitudes, derived subtotals) into
`PeriodFinancials`. The same logic generalises to any statement once a mapping
table is supplied.
"""

from __future__ import annotations

import openpyxl

from ..schema.financials import (
    BalanceSheet,
    CashFlow,
    CompanyFinancials,
    IncomeStatement,
    PeriodFinancials,
)


def _cell(ws, row, col):
    return ws.cell(row=row, column=col).value


def _find(ws, substring):
    for row in range(1, ws.max_row + 1):
        label = ws.cell(row=row, column=1).value
        if label and substring.lower() in str(label).lower():
            return _cell(ws, row, 2), _cell(ws, row, 3)
    return None, None


def _find_exact(ws, target):
    for row in range(1, ws.max_row + 1):
        label = ws.cell(row=row, column=1).value
        if label and str(label).strip().lower() == target.lower():
            return _cell(ws, row, 2), _cell(ws, row, 3)
    return None, None


def _to_float(v) -> float | None:
    if v is None:
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _abs(v) -> float | None:
    f = _to_float(v)
    return abs(f) if f is not None else None


def load_sc_workbook(path: str) -> CompanyFinancials:
    wb = openpyxl.load_workbook(path, data_only=True)
    pl = wb["I. Profit_Loss"]
    bs = wb["I. Balance_Sheet"]
    cf = wb["I. Cashflow"]
    rep = wb["O. Report"]

    periods = ["FY2022", "FY2023"]

    def get(sheet, sub):
        b, c = _find(sheet, sub)
        return _to_float(b), _to_float(c)

    is_periods = []
    for idx in range(2):
        rev = get(pl, "Revenue")[idx]
        cogs = get(pl, "Cost of sales")[idx]
        gp = get(pl, "Gross Profit")[idx]
        rnd = get(pl, "Research and development")[idx] or 0.0
        sga = get(pl, "Selling, general and administrative")[idx] or 0.0
        rest = get(pl, "Restructuring and other")[idx] or 0.0
        opex = rnd + sga + rest
        ebit = get(pl, "Income (loss) from operations")[idx]
        depr = _to_float(get(pl, "Depreciation & Amortisation")[idx]) or _to_float(get(pl, "Depreciation ")[idx])
        ebitda = _to_float(_find_exact(pl, "EBITDA")[idx])
        ie = _abs(get(pl, "Interest expense")[idx])
        ii = _to_float(get(pl, "Interest income")[idx])
        pbt = get(pl, "Profit Before Tax")[idx]
        tax = get(pl, "Tax")[idx]
        pat = get(pl, "Profit After Tax")[idx]
        is_periods.append(IncomeStatement(
            revenue=rev, cogs=cogs, gross_profit=gp, operating_expenses=opex,
            ebitda=ebitda, depreciation_amortization=depr, ebit=ebit,
            interest_expense=ie, interest_income=ii, pretax_income=pbt,
            tax_expense=tax, net_income=pat,
        ))

    bs_periods = []
    for idx in range(2):
        cash = get(bs, "Cash and cash equivalents")[idx]
        mkt = get(bs, "Short-term marketable securities")[idx]
        ar = get(bs, "Accounts receivable, net")[idx]
        inv = get(bs, "Inventories")[idx]
        ca = get(bs, "Total Current Assets")[idx]
        ta = get(bs, "Total Assets")[idx]
        ap = get(bs, "Accounts payable")[idx]
        cl = get(bs, "Total Current Liabilities")[idx]
        tl = get(bs, "Total Liabilities")[idx]
        std = get(bs, "Current portion of debt and finance leases")[idx]
        ltd = get(bs, "Debt and finance leases, net of current portion")[idx]
        td = None
        if std is not None or ltd is not None:
            td = (std or 0.0) + (ltd or 0.0)
        eq = get(bs, "Total Equity")[idx]
        re = get(bs, "Retained earnings (accumulated deficit)")[idx]
        bs_periods.append(BalanceSheet(
            cash_and_equivalents=cash, marketable_securities=mkt,
            accounts_receivable=ar, inventory=inv, current_assets=ca,
            total_assets=ta, accounts_payable=ap, current_liabilities=cl,
            total_liabilities=tl, short_term_debt=std, long_term_debt=ltd,
            total_debt=td, total_equity=eq, retained_earnings=re,
        ))

    cf_periods = []
    capex_labels = [
        "Purchases of property and equipment excluding finance leases, net of sales",
        "Purchases of solar energy systems, net of sales",
        "Purchases of digital assets",
        "Purchase of intangible assets",
    ]
    for idx in range(2):
        col = 2 + idx
        ocf = 0.0
        for r in range(6, 25):
            v = _to_float(_cell(cf, r, col))
            if v is not None:
                ocf += v
        icf = get(cf, "Net cash used in investing activities")[idx]
        capex = 0.0
        for lab in capex_labels:
            v = _to_float(get(cf, lab)[idx])
            if v is not None:
                capex += v
        capex = _abs(capex) if capex != 0 else None
        fcf = None
        if capex is not None:
            fcf = ocf - capex
        cf_periods.append(CashFlow(
            operating_cash_flow=ocf, capital_expenditures=capex,
            free_cash_flow=fcf, investing_cash_flow=icf,
        ))

    return CompanyFinancials(
        entity_name="Green Solutions Manufacturing Ltd",
        currency="USD (thousands)",
        periods=[
            PeriodFinancials(period=periods[0], income_statement=is_periods[0],
                             balance_sheet=bs_periods[0], cash_flow=cf_periods[0]),
            PeriodFinancials(period=periods[1], income_statement=is_periods[1],
                             balance_sheet=bs_periods[1], cash_flow=cf_periods[1]),
        ],
    )


def load_sc_benchmark(path: str) -> dict[str, dict[str, float]]:
    """Extract the expected ratios from the O. Report sheet for validation."""
    wb = openpyxl.load_workbook(path, data_only=True)
    rep = wb["O. Report"]
    out: dict[str, dict[str, float]] = {"FY2022": {}, "FY2023": {}}

    mapping = {
        "EBITDA Margin": "ebitda_margin",
        "Operating Margin": "operating_margin",
        "EBITDA Net Interest Cover (x)": "ebitda_net_interest_cover",
        "Current Ratio i.e. Current Assets / Current Liabilities (%)": "current_ratio",
        "Leverage Metric": "leverage_metric",
        "Net Leverage": "net_leverage",
        "Cash flow to capital expenditures (CF/CapEX)": "cf_to_capex",
    }
    for label, key in mapping.items():
        b, c = _find(rep, label)
        out["FY2022"][key] = float(b)
        out["FY2023"][key] = float(c)
    return out
