"""Tests for messy-excel improvements: number parsing, multi-period, labels, quality flags."""

from __future__ import annotations

import openpyxl
import pytest

from credit_agent.ingest.number_parser import parse_number
from credit_agent.ingest.generic_xlsx import (
    parse_generic_xlsx_multi,
    _is_period_header,
    _extract_year_label,
    _match_field,
)
from credit_agent.ingest.loader import ingest, ingest_file, _quality_flags


# ═══════════════════════════════════════════════════════════════════════════════
# Priority 2: Robust number parsing
# ═══════════════════════════════════════════════════════════════════════════════

class TestNumberParser:
    def test_plain_integer(self):
        assert parse_number("1234") == 1234.0

    def test_plain_float(self):
        assert parse_number("1234.56") == 1234.56

    def test_integer_input(self):
        assert parse_number(1234) == 1234.0

    def test_float_input(self):
        assert parse_number(1234.56) == 1234.56

    def test_none_input(self):
        assert parse_number(None) is None

    def test_empty_string(self):
        assert parse_number("") is None

    def test_comma_separated(self):
        assert parse_number("1,234,567") == 1234567.0

    def test_currency_dollar(self):
        assert parse_number("$1,234") == 1234.0

    def test_currency_pound(self):
        assert parse_number("\u00a31,234") == 1234.0

    def test_currency_euro(self):
        assert parse_number("\u20ac1,234") == 1234.0

    def test_parenthetical_negative(self):
        assert parse_number("(1,234)") == -1234.0

    def test_parenthetical_negative_no_comma(self):
        assert parse_number("(1234)") == -1234.0

    def test_leading_minus(self):
        assert parse_number("-1,234") == -1234.0

    def test_percentage(self):
        assert parse_number("12.3%") == 12.3

    def test_percentage_ratio_mode(self):
        assert parse_number("12.3%", pct_is_ratio=True) == pytest.approx(0.123)

    def test_parenthetical_percentage(self):
        assert parse_number("(12.3%)") == -12.3

    def test_em_dash_zero(self):
        assert parse_number("\u2014") == 0.0

    def test_en_dash_zero(self):
        assert parse_number("\u2013") == 0.0

    def test_hyphen_zero(self):
        assert parse_number("-") == 0.0

    def test_text_no_data(self):
        assert parse_number("n/a") is None
        assert parse_number("N/A") is None
        assert parse_number("nil") is None
        assert parse_number("NIL") is None

    def test_nan_returns_none(self):
        import math
        assert parse_number(float("nan")) is None

    def test_negative_with_currency(self):
        assert parse_number("($1,234)") == -1234.0

    def test_underscore_number(self):
        assert parse_number("1_234") == 1234.0

    def test_large_number(self):
        assert parse_number("1,234,567,890") == 1234567890.0


# ═══════════════════════════════════════════════════════════════════════════════
# Priority 1: Multi-period generic xlsx
# ═══════════════════════════════════════════════════════════════════════════════

class TestPeriodDetection:
    def test_is_period_header_year(self):
        assert _is_period_header(2023) is True
        assert _is_period_header("2023") is True

    def test_is_period_header_fy(self):
        assert _is_period_header("FY23") is True
        assert _is_period_header("FY2023") is True

    def test_is_period_header_date(self):
        assert _is_period_header("2023-12-31") is True
        assert _is_period_header("2023/12") is True

    def test_is_period_header_range(self):
        assert _is_period_header("23/24") is True
        assert _is_period_header("2022-2023") is True

    def test_is_period_header_not(self):
        assert _is_period_header("Revenue") is False
        assert _is_period_header("Total Assets") is False
        assert _is_period_header(None) is False
        assert _is_period_header("") is False

    def test_extract_year_full(self):
        assert _extract_year_label("2023") == "2023"
        assert _extract_year_label("2023-12-31") == "2023"
        assert _extract_year_label("FY2023") == "2023"

    def test_extract_year_short(self):
        assert _extract_year_label("FY23") == "2023"
        assert _extract_year_label("23/24") == "2024"
        assert _extract_year_label("22-23") == "2023"


class TestLabelMatching:
    def test_revenue_synonyms(self):
        assert _match_field("Revenue") == "revenue"
        assert _match_field("Net Revenue") == "revenue"
        assert _match_field("Total Revenue") == "revenue"
        assert _match_field("Net Sales") == "revenue"
        assert _match_field("Turnover") == "revenue"
        assert _match_field("Revenue from Operations") == "revenue"

    def test_cogs_synonyms(self):
        assert _match_field("Cost of Sales") == "cogs"
        assert _match_field("Cost of Revenue") == "cogs"
        assert _match_field("Cost of Goods Sold") == "cogs"

    def test_balance_sheet(self):
        assert _match_field("Total Assets") == "total_assets"
        assert _match_field("Shareholders Equity") == "total_equity"
        assert _match_field("Stockholders Equity") == "total_equity"
        assert _match_field("Trade Receivables") == "accounts_receivable"

    def test_cash_flow(self):
        assert _match_field("Cash Flow from Operating Activities") == "operating_cash_flow"
        assert _match_field("Net Cash from Operating Activities") == "operating_cash_flow"
        assert _match_field("Capital Expenditure") == "capital_expenditures"
        assert _match_field("Capex") == "capital_expenditures"

    def test_no_match(self):
        assert _match_field("Some Random Label") is None
        assert _match_field("") is None


class TestMultiPeriodGenericXlsx:
    def _make_xlsx(self, tmp_path, name, headers, rows):
        """Helper: create a generic xlsx with headers in row 1 and data rows."""
        path = tmp_path / name
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(headers)  # row 1: ["Label", "2021", "2022", "2023"]
        for row in rows:
            ws.append(row)
        wb.save(path)
        return str(path)

    def test_two_periods(self, tmp_path):
        path = self._make_xlsx(tmp_path, "test.xlsx",
            ["Line Item", "2022", "2023"],
            [
                ["Revenue", 1000, 1200],
                ["Cost of Sales", 600, 700],
                ["Total Assets", 5000, 5500],
                ["Total Equity", 3000, 3200],
            ],
        )
        periods = parse_generic_xlsx_multi(path, "2023")
        assert len(periods) == 2
        assert periods[0].period == "2022"
        assert periods[1].period == "2023"
        assert periods[0].income_statement.revenue == 1000
        assert periods[1].income_statement.revenue == 1200
        assert periods[0].balance_sheet.total_assets == 5000
        assert periods[1].balance_sheet.total_assets == 5500

    def test_three_periods(self, tmp_path):
        path = self._make_xlsx(tmp_path, "test3.xlsx",
            ["Item", "FY21", "FY22", "FY23"],
            [
                ["Revenue", 800, 1000, 1200],
                ["Net Income", 50, 80, 100],
                ["Total Assets", 4000, 5000, 5500],
            ],
        )
        periods = parse_generic_xlsx_multi(path, "2023")
        assert len(periods) == 3
        assert [p.period for p in periods] == ["2021", "2022", "2023"]
        assert periods[2].income_statement.revenue == 1200

    def test_single_period_fallback(self, tmp_path):
        path = self._make_xlsx(tmp_path, "single.xlsx",
            ["Item", "Value"],
            [
                ["Revenue", 1000],
                ["Total Assets", 5000],
            ],
        )
        periods = parse_generic_xlsx_multi(path, "2023")
        assert len(periods) == 1

    def test_parenthetical_negatives(self, tmp_path):
        path = self._make_xlsx(tmp_path, "neg.xlsx",
            ["Line Item", "2023"],
            [
                ["Revenue", 1000],
                ["Net Income", "(200)"],
                ["Total Equity", "(500)"],
            ],
        )
        periods = parse_generic_xlsx_multi(path, "2023")
        assert periods[0].income_statement.net_income == -200.0
        assert periods[0].balance_sheet.total_equity == -500.0

    def test_currency_values(self, tmp_path):
        path = self._make_xlsx(tmp_path, "cur.xlsx",
            ["Line Item", "2023"],
            [
                ["Revenue", "$1,234,567"],
                ["Cash and Cash Equivalents", "$50,000"],
            ],
        )
        periods = parse_generic_xlsx_multi(path, "2023")
        assert periods[0].income_statement.revenue == 1234567.0
        assert periods[0].balance_sheet.cash_and_equivalents == 50000.0

    def test_em_dash_as_zero(self, tmp_path):
        path = self._make_xlsx(tmp_path, "dash.xlsx",
            ["Line Item", "2023"],
            [
                ["Revenue", 1000],
                ["EBITDA", "\u2014"],
            ],
        )
        periods = parse_generic_xlsx_multi(path, "2023")
        assert periods[0].income_statement.revenue == 1000
        assert periods[0].income_statement.ebitda == 0.0

    def test_backward_compat_single_period(self, tmp_path):
        from credit_agent.ingest.generic_xlsx import parse_generic_xlsx
        path = self._make_xlsx(tmp_path, "compat.xlsx",
            ["Line Item", "2023"],
            [
                ["Revenue", 1000],
                ["Total Assets", 5000],
            ],
        )
        pf = parse_generic_xlsx(str(path), "2023")
        assert pf.period == "2023"
        assert pf.income_statement.revenue == 1000


# ═══════════════════════════════════════════════════════════════════════════════
# Priority 4: Data quality flags
# ═══════════════════════════════════════════════════════════════════════════════

class TestDataQualityFlags:
    def _make_period(self, year, revenue=None, net_income=None, total_assets=None,
                     total_equity=None, total_debt=None, ocf=None):
        from credit_agent.schema.financials import (
            BalanceSheet, CashFlow, IncomeStatement, PeriodFinancials,
        )
        return PeriodFinancials(
            period=str(year),
            income_statement=IncomeStatement(revenue=revenue, net_income=net_income),
            balance_sheet=BalanceSheet(
                total_assets=total_assets, total_equity=total_equity,
                total_debt=total_debt,
            ),
            cash_flow=CashFlow(operating_cash_flow=ocf),
        )

    def test_missing_critical_fields(self):
        p = self._make_period(2023)  # all None
        flags = _quality_flags([p])
        msg = " ".join(f.message for f in flags)
        assert "Revenue" in msg
        assert "Total Assets" in msg

    def test_negative_equity_warning(self):
        p = self._make_period(2023, total_equity=-500)
        flags = _quality_flags([p])
        assert any("negative total equity" in f.message for f in flags)

    def test_large_change_warning(self):
        p1 = self._make_period(2022, revenue=1000)
        p2 = self._make_period(2023, revenue=8000)
        flags = _quality_flags([p1, p2])
        assert any("700%" in f.message for f in flags)

    def test_no_warning_for_small_change(self):
        p1 = self._make_period(2022, revenue=1000)
        p2 = self._make_period(2023, revenue=1100)
        flags = _quality_flags([p1, p2])
        assert not any("Revenue changed" in f.message for f in flags)

    def test_empty_periods(self):
        flags = _quality_flags([])
        assert any("No periods" in f.message for f in flags)


# ═══════════════════════════════════════════════════════════════════════════════
# Integration: ingest() with messy generic xlsx
# ═══════════════════════════════════════════════════════════════════════════════

class TestMessyIngestIntegration:
    def test_generic_xlsx_multi_period_via_ingest(self, tmp_path):
        path = tmp_path / "messy.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Line Item", "FY21", "FY22", "FY23"])
        ws.append(["Revenue", 800, 1000, 1200])
        ws.append(["Cost of Sales", 500, 600, 700])
        ws.append(["Total Assets", 4000, 5000, 5500])
        ws.append(["Total Equity", 2500, 3000, 3200])
        ws.append(["Total Debt", 1000, 1200, 1100])
        wb.save(path)

        res = ingest([{"path": str(path), "year": 2023, "entity": "MessyCo"}])
        assert res.entity_name == "MessyCo"
        assert len(res.periods) == 3
        assert [p.period for p in res.periods] == ["2021", "2022", "2023"]
        assert res.periods[2].income_statement.revenue == 1200
        assert any("warning" in f.level for f in res.flags)

    def test_sc_workbook_still_works(self):
        SC = "data/raw/Task 1 Example Answer - Financial Reporting Tool.xlsx"
        res = ingest([{"path": SC, "year": 2023, "entity": "GS"}])
        assert res.entity_name == "GS"
        assert [p.period for p in res.periods] == ["2022", "2023"]
        assert res.currency == "USD (thousands)"


# ── Merged cell handling ─────────────────────────────────────────────────────

class TestMergedCells:
    def test_merged_label_is_read(self, tmp_path):
        from credit_agent.ingest.generic_xlsx import parse_generic_xlsx_multi
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.merge_cells("A1:A3")
        ws["A1"] = "Revenue"
        ws["B1"] = 100
        ws["B2"] = 200
        ws["B3"] = 300
        ws["A4"] = "COGS"
        ws["B4"] = 50
        ws["A5"] = "Net Income"
        ws["B5"] = 30

        path = tmp_path / "merged.xlsx"
        wb.save(str(path))
        periods = parse_generic_xlsx_multi(str(path), "2023")
        assert len(periods) == 1
        assert periods[0].income_statement.revenue == 100.0

    def test_merged_period_header(self, tmp_path):
        from credit_agent.ingest.generic_xlsx import parse_generic_xlsx_multi
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        # Merge cells in row 1 for a year header spanning 2 cols
        ws.merge_cells("B1:C1")
        ws["A1"] = ""
        ws["B1"] = "2023"
        ws["A2"] = "Revenue"
        ws["B2"] = 100
        ws["A3"] = "COGS"
        ws["B3"] = 50

        path = tmp_path / "merged_header.xlsx"
        wb.save(str(path))
        periods = parse_generic_xlsx_multi(str(path), "2023")
        assert len(periods) >= 1
        assert periods[0].income_statement.revenue == 100.0


# ── Hidden rows/columns ─────────────────────────────────────────────────────

class TestHiddenRowsCols:
    def test_hidden_row_is_skipped(self, tmp_path):
        from credit_agent.ingest.generic_xlsx import parse_generic_xlsx_multi
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Revenue"
        ws["B1"] = 100
        ws["A2"] = "COGS"
        ws["B2"] = 50
        ws.row_dimensions[2].hidden = True  # hide COGS row

        path = tmp_path / "hidden.xlsx"
        wb.save(str(path))
        periods = parse_generic_xlsx_multi(str(path), "2023")
        assert periods[0].income_statement.revenue == 100.0
        assert periods[0].income_statement.cogs is None

    def test_hidden_col_is_skipped(self, tmp_path):
        from credit_agent.ingest.generic_xlsx import parse_generic_xlsx_multi
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Revenue"
        ws["B1"] = 100
        ws["C1"] = 200
        ws.column_dimensions["C"].hidden = True  # hide col C

        path = tmp_path / "hidden_col.xlsx"
        wb.save(str(path))
        periods = parse_generic_xlsx_multi(str(path), "2023")
        assert len(periods) == 1
        # Only one data column visible (col B), so only one period
        assert periods[0].income_statement.revenue == 100.0


# ── Vertical layout ─────────────────────────────────────────────────────────

class TestVerticalLayout:
    def test_vertical_years_as_rows(self, tmp_path):
        from credit_agent.ingest.generic_xlsx import parse_generic_xlsx_multi
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        # Vertical layout: labels in row 1, years as rows
        ws["A1"] = ""
        ws["B1"] = "Revenue"
        ws["C1"] = "COGS"
        ws["A2"] = "2022"
        ws["B2"] = 1000
        ws["C2"] = 400
        ws["A3"] = "2023"
        ws["B3"] = 1200
        ws["C3"] = 500

        path = tmp_path / "vertical.xlsx"
        wb.save(str(path))
        periods = parse_generic_xlsx_multi(str(path), "2023")
        assert len(periods) == 2
        by_yr = {p.period: p for p in periods}
        assert by_yr["2022"].income_statement.revenue == 1000.0
        assert by_yr["2023"].income_statement.revenue == 1200.0
        assert by_yr["2022"].income_statement.cogs == 400.0
        assert by_yr["2023"].income_statement.cogs == 500.0


# ── Multi-sheet detection ───────────────────────────────────────────────────

class TestMultiSheetDetection:
    def test_classify_sheet_names(self):
        from credit_agent.ingest.generic_xlsx import _classify_sheet
        assert _classify_sheet("Income Statement") == "income_statement"
        assert _classify_sheet("Balance Sheet") == "balance_sheet"
        assert _classify_sheet("Cash Flow") == "cash_flow"
        assert _classify_sheet("P&L") == "income_statement"
        assert _classify_sheet("Profit & Loss") == "income_statement"
        assert _classify_sheet("Statement of Financial Position") == "balance_sheet"
        assert _classify_sheet("Cashflows") == "cash_flow"
        assert _classify_sheet("Random Data") is None

    def test_describe_workbook(self):
        from credit_agent.ingest.generic_xlsx import describe_workbook
        SC = "data/raw/Task 1 Example Answer - Financial Reporting Tool.xlsx"
        info = describe_workbook(SC)
        assert "sheet_names" in info
        assert len(info["sheet_names"]) > 0
        # Our SC workbook has known sheets
        types = info["sheet_types"]
        assert any(v == "income_statement" for v in types.values()), (
            f"Expected IS in types, got {types}"
        )
        assert any(v == "balance_sheet" for v in types.values()), (
            f"Expected BS in types, got {types}"
        )
        assert any(v == "cash_flow" for v in types.values()), (
            f"Expected CF in types, got {types}"
        )

    def test_multi_sheet_xlsx_combines_data(self, tmp_path):
        from credit_agent.ingest.generic_xlsx import parse_generic_xlsx_multi
        from openpyxl import Workbook

        wb = Workbook()
        # Sheet 1: Income Statement
        ws_is = wb.active
        ws_is.title = "Income Statement"
        ws_is["A1"] = "Revenue"
        ws_is["B1"] = 1000
        ws_is["A2"] = "COGS"
        ws_is["B2"] = 400
        ws_is["A3"] = "Net Income"
        ws_is["B3"] = 200

        # Sheet 2: Balance Sheet
        ws_bs = wb.create_sheet("Balance Sheet")
        ws_bs["A1"] = "Cash"
        ws_bs["B1"] = 150
        ws_bs["A2"] = "Total Assets"
        ws_bs["B2"] = 3000
        ws_bs["A3"] = "Total Equity"
        ws_bs["B3"] = 1800

        # Sheet 3: Cash Flow
        ws_cf = wb.create_sheet("Cash Flow")
        ws_cf["A1"] = "Operating Cash Flow"
        ws_cf["B1"] = 350

        path = tmp_path / "multi_sheet.xlsx"
        wb.save(str(path))
        periods = parse_generic_xlsx_multi(str(path), "2023")
        assert len(periods) == 1
        p = periods[0]
        assert p.income_statement.revenue == 1000.0
        assert p.income_statement.cogs == 400.0
        assert p.balance_sheet.total_assets == 3000.0
        assert p.balance_sheet.total_equity == 1800.0
        assert p.cash_flow.operating_cash_flow == 350.0
